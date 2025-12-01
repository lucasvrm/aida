from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from app.core.config import settings
from app.core.br_formats import normalize_header
from app.core.utils import normalize_whitespace
from app.models.payload import ConsolidatedPayload
from app.template.specs import (
    RECEBIVEIS, TIPOLOGIA, LANDBANK, ENDIVIDAMENTO, VIABILIDADE, PROJETO_CELLS, TableSpec
)

@dataclass
class WriteResult:
    out_path: str

def _load_kv_spec() -> dict[str, Any]:
    import json
    p = Path(settings.KV_SPEC_PATH)
    return json.loads(p.read_text(encoding="utf-8"))

def _clear_table(ws, spec: TableSpec, max_rows: int) -> None:
    start = spec.data_start_row
    end = start + max_rows - 1
    for (col_letter, _name) in spec.columns:
        for r in range(start, end + 1):
            ws[f"{col_letter}{r}"].value = None

def _write_table(ws, spec: TableSpec, rows: list[dict[str, Any]], max_rows: int) -> None:
    if len(rows) > max_rows:
        rows = rows[:max_rows]
    for i, row in enumerate(rows):
        r = spec.data_start_row + i
        for (col_letter, _name) in spec.columns:
            if col_letter in row:
                v = row[col_letter]
                if isinstance(v, str):
                    v = normalize_whitespace(v)
                ws[f"{col_letter}{r}"].value = v

def _write_geral_kv(wb, data: dict[str, Any]) -> None:
    ws = wb["Geral"]
    spec = _load_kv_spec()
    by_norm = spec.get("by_label_norm", {})
    for label, value in (data or {}).items():
        ln = normalize_header(label)
        item = by_norm.get(ln)
        if not item:
            for k, it in by_norm.items():
                if ln and ln in k:
                    item = it
                    break
        if not item:
            continue
        ws[item["value_cell"]].value = value

def _write_projeto_cells(wb, data: dict[str, Any]) -> None:
    ws = wb["Projeto"]
    for key, cell in PROJETO_CELLS.items():
        if not data or key not in data:
            continue
        ws[cell].value = data[key]

def write_filled_xlsx(payload: ConsolidatedPayload, project_name: str, out_path: str) -> WriteResult:
    wb = openpyxl.load_workbook(settings.TEMPLATE_PATH)

    if "Geral" in wb.sheetnames:
        _write_geral_kv(wb, payload.Geral.data)

    if "Projeto" in wb.sheetnames:
        _write_projeto_cells(wb, payload.Projeto.data)

    table_specs = [
        (RECEBIVEIS, payload.Recebíveis.rows),
        (TIPOLOGIA, payload.Tipologia.rows),
        (LANDBANK, payload.Landbank.rows),
        (ENDIVIDAMENTO, payload.Endividamento.rows),
        (VIABILIDADE, payload.Viabilidade_Financeira.rows),
    ]
    for spec, rows in table_specs:
        if spec.sheet not in wb.sheetnames:
            continue
        ws = wb[spec.sheet]
        _clear_table(ws, spec, settings.MAX_TABLE_ROWS)
        _write_table(ws, spec, rows or [], settings.MAX_TABLE_ROWS)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return WriteResult(out_path=out_path)
