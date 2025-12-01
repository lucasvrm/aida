from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl

from app.core.config import settings
from app.core.br_formats import normalize_header

def generate_kv_spec_from_template(template_path: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(template_path)
    if "Geral" not in wb.sheetnames:
        raise RuntimeError("Template sem aba 'Geral'.")

    ws = wb["Geral"]
    spec: dict[str, Any] = {
        "by_label_norm": {},
        "pairs": [],
    }

    def scan(label_col: str, value_col: str, row_min: int = 1, row_max: int = 250):
        for r in range(row_min, row_max + 1):
            label_cell = f"{label_col}{r}"
            value_cell = f"{value_col}{r}"
            label = ws[label_cell].value
            if label is None:
                continue
            label_str = str(label).strip()
            if not label_str:
                continue
            label_norm = normalize_header(label_str)
            item = {"label": label_str, "label_cell": label_cell, "value_cell": value_cell}
            spec["pairs"].append(item)
            spec["by_label_norm"].setdefault(label_norm, item)

    scan("B", "C")
    scan("E", "F")
    return spec

def ensure_kv_spec() -> None:
    spec_path = Path(settings.KV_SPEC_PATH)
    if spec_path.exists():
        try:
            data = json.loads(spec_path.read_text(encoding="utf-8"))
            if isinstance(data, dict) and data.get("pairs") and data.get("by_label_norm"):
                return
        except Exception:
            pass

    spec = generate_kv_spec_from_template(settings.TEMPLATE_PATH)
    spec_path.parent.mkdir(parents=True, exist_ok=True)
    spec_path.write_text(json.dumps(spec, ensure_ascii=False, indent=2), encoding="utf-8")
